#!/usr/bin/env python3
"""
生成估值计算模板和分析Checklist
基于唐朝投资思想
"""

import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError:
    print("需要安装openpyxl: pip install openpyxl")
    exit(1)


def create_valuation_template():
    """创建估值计算模板"""

    wb = openpyxl.Workbook()

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14)
    input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    calc_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")

    # ========== Sheet 1: DCF估值模型 ==========
    ws1 = wb.active
    ws1.title = "DCF估值模型"

    ws1['A1'] = "自由现金流折现（DCF）估值模型"
    ws1['A1'].font = title_font
    ws1.merge_cells('A1:F1')

    ws1['A3'] = "基本信息"
    ws1['A3'].font = Font(bold=True)

    info_data = [
        ["项目", "数值", "说明"],
        ["公司名称", "", "请输入公司名称"],
        ["股票代码", "", "如：600519"],
        ["分析日期", "", datetime.now().strftime("%Y-%m-%d")],
        ["当前股价", "", "元"],
        ["总股本", "", "亿股"],
        ["当前市值", '=B6*B7', "亿元（自动计算）"],
    ]

    for row_idx, row_data in enumerate(info_data, start=4):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 4:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            if col_idx == 2 and row_idx > 4 and '=' not in str(value):
                cell.fill = input_fill
            if '=' in str(value):
                cell.fill = calc_fill
            cell.border = thin_border

    ws1['A13'] = "核心假设（请填写黄色区域）"
    ws1['A13'].font = Font(bold=True)

    dcf_data = [
        ["项目", "数值", "单位", "说明"],
        ["当年净利润", "", "亿元", "最近年报净利润"],
        ["分红率", 0.5, "", "历史分红率，如50%填0.5"],
        ["前5年增长率", 0.15, "", "预计增长率，如15%填0.15"],
        ["后5年增长率", 0.10, "", "预计增长率放缓"],
        ["永续增长率", 0.04, "", "通常3%-5%"],
        ["折现率", 0.10, "", "无风险利率+风险溢价，如10%"],
    ]

    for row_idx, row_data in enumerate(dcf_data, start=14):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 14:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            if col_idx == 2 and row_idx > 14:
                cell.fill = input_fill
            cell.border = thin_border

    ws1['A22'] = "自由现金流预测"
    ws1['A22'].font = Font(bold=True)

    fcf_header = ["年份", "净利润(亿元)", "自由现金流(亿元)", "折现因子", "现值(亿元)"]
    for col_idx, value in enumerate(fcf_header, start=1):
        cell = ws1.cell(row=23, column=col_idx, value=value)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for year in range(1, 11):
        row_idx = 23 + year
        ws1.cell(row=row_idx, column=1, value=f"第{year}年").border = thin_border
        if year <= 5:
            growth_rate = "B18"
        else:
            growth_rate = "B19"

        if year == 1:
            profit_formula = f'=B15*(1-B16)'
        elif year <= 5:
            profit_formula = f'=B{23+year-1}*(1+$B$18)'
        else:
            profit_formula = f'=B{23+year-1}*(1+$B$19)'

        ws1.cell(row=row_idx, column=2, value=f'=B15*(1-B16)*POWER(1+IF({year}<=5,$B$18,$B$19),{year})').border = thin_border
        ws1.cell(row=row_idx, column=2).fill = calc_fill

        ws1.cell(row=row_idx, column=3, value=f'=B{row_idx}').border = thin_border
        ws1.cell(row=row_idx, column=3).fill = calc_fill

        ws1.cell(row=row_idx, column=4, value=f'=1/POWER(1+$B$20,{year})').border = thin_border
        ws1.cell(row=row_idx, column=4).fill = calc_fill

        ws1.cell(row=row_idx, column=5, value=f'=C{row_idx}*D{row_idx}').border = thin_border
        ws1.cell(row=row_idx, column=5).fill = calc_fill

    row_idx = 34
    ws1.cell(row=row_idx, column=1, value="前10年现值合计").border = thin_border
    ws1.cell(row=row_idx, column=5, value="=SUM(E24:E33)").border = thin_border
    ws1.cell(row=row_idx, column=5).fill = calc_fill

    row_idx = 35
    ws1.cell(row=row_idx, column=1, value="第10年末终值").border = thin_border
    ws1.cell(row=row_idx, column=5, value="=C33*(1+B21)/(B20-B21)").border = thin_border
    ws1.cell(row=row_idx, column=5).fill = calc_fill

    row_idx = 36
    ws1.cell(row=row_idx, column=1, value="终值折现").border = thin_border
    ws1.cell(row=row_idx, column=5, value="=E35*D33").border = thin_border
    ws1.cell(row=row_idx, column=5).fill = calc_fill

    row_idx = 37
    ws1.cell(row=row_idx, column=1, value="内在价值（总）").border = thin_border
    ws1.cell(row=row_idx, column=5, value="=E34+E36").border = thin_border
    ws1.cell(row=row_idx, column=5).fill = calc_fill

    row_idx = 38
    ws1.cell(row=row_idx, column=1, value="每股内在价值").border = thin_border
    ws1.cell(row=row_idx, column=5, value="=E37/B7*10000").border = thin_border
    ws1.cell(row=row_idx, column=5).fill = calc_fill

    row_idx = 39
    ws1.cell(row=row_idx, column=1, value="5折买入价").border = thin_border
    ws1.cell(row=row_idx, column=5, value="=E38*0.5").border = thin_border
    ws1.cell(row=row_idx, column=5).fill = calc_fill

    row_idx = 40
    ws1.cell(row=row_idx, column=1, value="150%卖出价").border = thin_border
    ws1.cell(row=row_idx, column=5, value="=E38*1.5").border = thin_border
    ws1.cell(row=row_idx, column=5).fill = calc_fill

    for col in ['A', 'B', 'C', 'D', 'E']:
        ws1.column_dimensions[col].width = 18

    # ========== Sheet 2: 财务比率 ==========
    ws2 = wb.create_sheet("财务比率")

    ws2['A1'] = "关键财务比率分析"
    ws2['A1'].font = title_font
    ws2.merge_cells('A1:C1')

    ratio_data = [
        ["比率", "数值", "标准值/行业平均"],
        ["毛利率(%)", "", ">40%优秀"],
        ["净利率(%)", "", ">20%优秀"],
        ["ROE(%)", "", ">20%优秀"],
        ["ROA(%)", "", ">10%良好"],
        ["资产负债率(%)", "", "<50%安全"],
        ["流动比率", "", ">2安全"],
        ["速动比率", "", ">1安全"],
        ["存货周转天数", "", "越低越好"],
        ["应收账款周转天数", "", "越低越好"],
        ["经营现金流/净利润", "", ">1健康"],
        ["分红率(%)", "", "30-50%合理"],
    ]

    for row_idx, row_data in enumerate(ratio_data, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 3:
                cell.fill = header_fill
                cell.font = header_font
            if col_idx == 2 and row_idx > 3:
                cell.fill = input_fill
            cell.border = thin_border

    for col in ['A', 'B', 'C']:
        ws2.column_dimensions[col].width = 25

    # ========== Sheet 3: 利润质量检查 ==========
    ws3 = wb.create_sheet("利润质量检查")

    ws3['A1'] = "利润质量检查清单"
    ws3['A1'].font = title_font
    ws3.merge_cells('A1:D1')

    check_data = [
        ["检查项目", "数值", "判断标准", "结论"],
        ["经营现金流/净利润", "", ">1.0为健康", ""],
        ["销售现金/营业收入", "", ">1.17为健康", ""],
        ["扣非净利润/净利润", "", ">0.8为健康", ""],
        ["应收账款增长率", "", "<=营收增长率", ""],
        ["存货增长率", "", "<=营收增长率", ""],
        ["其他应收款/总资产", "", "<0.05正常", ""],
        ["商誉/总资产", "", "<0.1正常", ""],
    ]

    for row_idx, row_data in enumerate(check_data, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 3:
                cell.fill = header_fill
                cell.font = header_font
            if col_idx == 2 and row_idx > 3:
                cell.fill = input_fill
            cell.border = thin_border

    for col in ['A', 'B', 'C', 'D']:
        ws3.column_dimensions[col].width = 25

    # ========== Sheet 4: 投资决策 ==========
    ws4 = wb.create_sheet("投资决策")

    ws4['A1'] = "投资决策汇总"
    ws4['A1'].font = title_font
    ws4.merge_cells('A1:C1')

    decision_data = [
        ["项目", "结果", "备注"],
        ["当前股价", "", ""],
        ["每股内在价值", "", ""],
        ["5折买入价", "", ""],
        ["150%卖出价", "", ""],
        ["安全边际", '=IF(B3>0,(B3-B2)/B3,"")', "(内在价值-当前价)/内在价值"],
        ["买入建议", '=IF(B6>0.5,"强烈建议买入",IF(B6>0.3,"可以买入","等待"))', ""],
    ]

    for row_idx, row_data in enumerate(decision_data, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 3:
                cell.fill = header_fill
                cell.font = header_font
            if col_idx == 2 and row_idx > 3 and '=' not in str(value):
                cell.fill = input_fill
            if '=' in str(value):
                cell.fill = calc_fill
            cell.border = thin_border

    for col in ['A', 'B', 'C']:
        ws4.column_dimensions[col].width = 20

    return wb


def create_checklist_workbook():
    """创建分析Checklist"""

    wb = openpyxl.Workbook()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=14)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ========== Sheet 1: 商业模式Checklist ==========
    ws1 = wb.active
    ws1.title = "商业模式"

    ws1['A1'] = "商业模式分析Checklist"
    ws1['A1'].font = title_font
    ws1.merge_cells('A1:C1')

    items = [
        ["分析项目", "完成情况", "备注"],
        ["1. 核心业务是否一句话能说清楚？", "", ""],
        ["2. 主要收入来源是否明确？", "", ""],
        ["3. 毛利率>40%吗？", "", ""],
        ["4. 净利率>15%吗？", "", ""],
        ["5. ROE>15%吗？", "", ""],
        ["6. 产品是高端/普通/低端？", "", ""],
        ["7. 技术迭代速度如何？", "", ""],
        ["8. 客户是个人还是企业？", "", ""],
        ["9. 是先款后货还是先货后款？", "", ""],
        ["10. 应收账款多吗？", "", ""],
        ["11. 对上游付款方式如何？", "", ""],
        ["12. 波特五力情况如何？", "", ""],
        ["13. 是周期行业还是非周期行业？", "", ""],
    ]

    for row_idx, row_data in enumerate(items, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 3:
                cell.fill = header_fill
                cell.font = header_font
            cell.border = thin_border

    ws1.column_dimensions['A'].width = 45
    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 30

    # ========== Sheet 2: 护城河Checklist ==========
    ws2 = wb.create_sheet("护城河")

    ws2['A1'] = "护城河分析Checklist"
    ws2['A1'].font = title_font
    ws2.merge_cells('A1:C1')

    moat_items = [
        ["分析项目", "是/否", "说明"],
        ["1. 有明显的品牌优势吗？", "", ""],
        ["2. 消费者愿意为品牌支付溢价吗？", "", ""],
        ["3. 有网络效应吗？", "", ""],
        ["4. 有成本优势吗？", "", ""],
        ["5. 有专利或技术壁垒吗？", "", ""],
        ["6. 有牌照或监管壁垒吗？", "", ""],
        ["7. 有渠道优势吗？", "", ""],
        ["8. 有地理或资源优势吗？", "", ""],
        ["9. 护城河宽还是窄？", "", ""],
        ["10. 护城河可持续吗？", "", ""],
    ]

    for row_idx, row_data in enumerate(moat_items, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 3:
                cell.fill = header_fill
                cell.font = header_font
            cell.border = thin_border

    ws2.column_dimensions['A'].width = 40
    ws2.column_dimensions['B'].width = 10
    ws2.column_dimensions['C'].width = 40

    # ========== Sheet 3: 财报Checklist ==========
    ws3 = wb.create_sheet("财报检查")

    ws3['A1'] = "财报检查Checklist"
    ws3['A1'].font = title_font
    ws3.merge_cells('A1:D1')

    finance_items = [
        ["检查项目", "数值", "标准", "通过/不通过"],
        ["经营现金流 ≥ 净利润", "", "是", ""],
        ["销售现金 ≥ 营收×1.17", "", "是", ""],
        ["扣非净利润/净利润 ≥ 0.8", "", "是", ""],
        ["应收增长率 ≤ 营收增长率", "", "是", ""],
        ["存货增长率 ≤ 营收增长率", "", "是", ""],
        ["其他应收款/总资产 ≤ 0.05", "", "是", ""],
        ["商誉/总资产 ≤ 0.1", "", "是", ""],
        ["货币资金 ≥ 有息负债", "", "是", ""],
        ["流动比率 ≥ 1.5", "", "是", ""],
        ["分红率 ≥ 20%", "", "是", ""],
    ]

    for row_idx, row_data in enumerate(finance_items, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 3:
                cell.fill = header_fill
                cell.font = header_font
            cell.border = thin_border

    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 15
    ws3.column_dimensions['C'].width = 10
    ws3.column_dimensions['D'].width = 12

    # ========== Sheet 4: 投资决策 ==========
    ws4 = wb.create_sheet("投资决策")

    ws4['A1'] = "投资决策最终Checklist"
    ws4['A1'].font = title_font
    ws4.merge_cells('A1:C1')

    decision_items = [
        ["项目", "结论", "备注"],
        ["商业模式优秀吗？", "", ""],
        ["有护城河吗？", "", ""],
        ["管理层可信吗？", "", ""],
        ["利润是真的吗？", "", ""],
        ["利润可持续吗？", "", ""],
        ["有安全边际吗？>50%？", "", ""],
        ["愿意持有3年以上吗？", "", ""],
        ["敢重仓吗？", "", ""],
        ["综合结论：买入/持有/卖出？", "", ""],
    ]

    for row_idx, row_data in enumerate(decision_items, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 3:
                cell.fill = header_fill
                cell.font = header_font
            cell.border = thin_border

    ws4.column_dimensions['A'].width = 30
    ws4.column_dimensions['B'].width = 20
    ws4.column_dimensions['C'].width = 40

    return wb


def main():
    assets_dir = os.path.dirname(os.path.abspath(__file__))

    print("生成估值模板...")
    wb1 = create_valuation_template()
    template_path = os.path.join(assets_dir, "valuation_template.xlsx")
    wb1.save(template_path)
    print(f"已保存: {template_path}")

    print("生成分析Checklist...")
    wb2 = create_checklist_workbook()
    checklist_path = os.path.join(assets_dir, "analysis_checklist.xlsx")
    wb2.save(checklist_path)
    print(f"已保存: {checklist_path}")

    print("\n完成！两个Excel文件已生成。")
    print("\n文件说明:")
    print("1. valuation_template.xlsx - 估值计算模板")
    print("   - DCF估值模型: 自由现金流折现计算")
    print("   - 财务比率: 关键比率跟踪")
    print("   - 利润质量检查: 利润真实性检查")
    print("   - 投资决策: 买入/卖出点计算")
    print("\n2. analysis_checklist.xlsx - 企业分析Checklist")
    print("   - 商业模式: 商业模式分析检查项")
    print("   - 护城河: 竞争优势分析检查项")
    print("   - 财报检查: 关键财务指标检查")
    print("   - 投资决策: 最终投资决策清单")


if __name__ == "__main__":
    main()
